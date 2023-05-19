import matplotlib
import matplotlib.pyplot as pl
matplotlib.use("TKAgg")
import numpy as np
from matplotlib.backends.backend_tkagg import FigureCanvasTkAgg
from matplotlib.figure import Figure


import openpyxl 
import os
import tkinter as tki
from tkinter import *
from tkinter.font import Font
from tkinter import messagebox

import mysql.connector
import PIL
from PIL import Image,ImageTk
#Welcome Page
top=tki.Tk()
img=Image.open("654.png")
img=img.resize((1365,720),Image.ANTIALIAS)
photo=ImageTk.PhotoImage(img)
label=tki.Label(top,image=photo)
label.grid()


top.geometry("{0}x{1}+0+0".format(top.winfo_screenwidth(), top.winfo_screenheight()))
top.title("Kendriya vidyalaya")
top.iconbitmap(r'Icewind Dale_1.ico')

var=StringVar()
L1=Label(top,textvariable=var,relief=RAISED,width=17,height=1,bd=0,bg="LavenderBlush",\
         activebackground="#fff",activeforeground="#42f498",fg="black")
var.set("Welcome to KV AMC")
myfont=Font(family="Impact",size=60)
L1.configure(font=myfont)
L1.place(x=330,y=20)
#end of text


#image
canvas=Canvas(top,height=140,width=150)
canvas.place(x=600,y=250)
img=PhotoImage(file="kvs-logo.gif")
canvas.create_image(1, 1, anchor=NW, image=img)
#image closed


#entry
password=Label(top,text="Password").place(x=570,y=505)
v1=StringVar()
E1=Entry(top,textvariable=v1,bd=5,show="*").place(x=700,y=500)
#entry closed

#creating second window
def nxt():
    root=tki.Tk()
    '''
    canvas=Canvas(root,width=300,height=160)
    image=ImageTk.PhotoImage(Image.open("pp.jpg"))
    canvas.create_image(0,0,anchor=NW,image=image)
    canvas.pack()
    '''
    root.geometry("{0}x{1}+0+0".format(root.winfo_screenwidth(), root.winfo_screenheight()))
    root.title("Kendriya vidyalaya")
    L2=Label(root,text="Student details")
    font1=Font(family="Impact",size=60)
    L2.configure(font=font1)
    L2.grid(row=0,column=3)
    

    #primary image
    canvase2=Canvas(root,height=300,width=300,relief=RAISED)
    canvase2.grid(row=3,column=1)
    photo2=PhotoImage(file="school-kids.gif")
    canvase2.create_image(100,100,anchor=NW,image=photo2)

    #secondary image
    canvase=Canvas(root,height=300,width=300)
    canvase.grid(row=3,column=4)
    photo=PhotoImage(file="secondary-schools.gif")
    canvase.create_image(100,100,anchor=NW,image=photo)

    #functions for buttons

    def primary():
        def back():
            sut.destroy()
            nxt()
            
        root.destroy()
        sut=tki.Tk()

        canvas=Canvas(sut,width=300,height=160)
        image=ImageTk.PhotoImage(Image.open("pp.jpg"))
        canvas.create_image(0,0,anchor=NW,image=image)
        canvas.pack()

        img=Image.open("vnbv.jpg")
        sut.geometry("{0}x{1}+0+0".format(sut.winfo_screenwidth(), sut.winfo_screenheight()))
        sut.title("Kendriya vidyalaya")

        b=Label(sut,text='PRIMARY SECTION',font='arial 14 bold')
        myfont=Font(family="Impact",size=30)
        b.configure(font=myfont)
        b.place(x=460,y=0)



        button1=Button(sut,text="_          CLASS I         _",width=18,height=2,bd=0,bg="DodgerBlue",\
                       activebackground="#fff",activeforeground="#42f498",fg="#fff")
        button1.place(x=550,y=150)
        button2=Button(sut,text="_          CLASS II        _",width=18,height=2,bd=0,bg="DodgerBlue",\
                       activebackground="#fff",activeforeground="#42f498",fg="#fff")
        button2.place(x=550,y=200) 
        button3=Button(sut,text="_         CLASS III        _",width=18,height=2,bd=0,bg="DodgerBlue",\
                       activebackground="#fff",activeforeground="#42f498",fg="#fff")
        button3.place(x=550,y=250)
        button4=Button(sut,text="_          CLASS IV        _",width=18,height=2,bd=0,bg="DodgerBlue",\
                       activebackground="#fff",activeforeground="#42f498",fg="#fff")
        button4.place(x=550,y=300)
        button5=Button(sut,text="_          CLASS V         _",width=18,height=2,bd=0,bg="DodgerBlue",\
                       activebackground="#fff",activeforeground="#42f498",fg="#fff")
        button5.place(x=550,y=350)       
        button6=Button(sut,text="Back",command=back,width=18,height=2,bd=0,bg="DimGray",\
                       activebackground="#fff",activeforeground="#42f498",fg="#fff")
        button6.place(x=550,y=400)
        

        
    def secondary():
        def back1():
            nut.destroy()
            nxt()
        root.destroy()
        nut=tki.Tk()
        nut.geometry("{0}x{1}+0+0".format(nut.winfo_screenwidth(), nut.winfo_screenheight()))
        nut.title("Kendriya vidyalaya")
        b=Label(nut,text='SECONDARY SECTION')
        myfont=Font(family="Impact",size=30)
        b.configure(font=myfont)
        b.place(x=460,y=0)
        button1=Button(nut,text="_           CLASS VI          _",width=18,height=2,bd=0,bg="DodgerBlue",\
                       activebackground="#fff",activeforeground="#42f498",fg="#fff")
        button1.place(x=550,y=150)
        button2=Button(nut,text="_           CLASS VII         _",width=18,height=2,bd=0,bg="DodgerBlue",\
                       activebackground="#fff",activeforeground="#42f498",fg="#fff")
        button2.place(x=550,y=200)
        button3=Button(nut,text="_           CLASS VIII        _",width=18,height=2,bd=0,bg="DodgerBlue",\
                       activebackground="#fff",activeforeground="#42f498",fg="#fff")
        button3.place(x=550,y=250)
        button4=Button(nut,text="_           CLASS IX          _",width=18,height=2,bd=0,bg="DodgerBlue",\
                       activebackground="#fff",activeforeground="#42f498",fg="#fff")
        button4.place(x=550,y=300) 
        button5=Button(nut,text="_           CLASS X           _",width=18,height=2,bd=0,bg="DodgerBlue",\
                       activebackground="#fff",activeforeground="#42f498",fg="#fff")
        button5.place(x=550,y=350)
        button6=Button(nut,text="_           CLASS XI          _",width=18,height=2,bd=0,bg="DodgerBlue",\
                       activebackground="#fff",activeforeground="#42f498",fg="#fff")
        button6.place(x=550,y=400)
        button7=Button(nut,text="Back",command=back1,width=18,height=2,bd=0,bg="DimGray",\
                       activebackground="#fff",activeforeground="#42f498",fg="#fff")
        button7.place(x=550,y=550)

        #new window for class 12th           
        def class12():
            nut.destroy()
            cut=tki.Tk()
            cut.geometry("{0}x{1}+0+0".format(cut.winfo_screenwidth(), cut.winfo_screenheight()))
            cut.title("Kendriya vidyalaya")
            b=Label(cut,text='Class 12th',font='arial 14 bold')
            myfont=Font(family="Impact",size=30)
            b.configure(font=myfont)
            b.place(x=525,y=0)
            #sections
####################################################################################################################################################################            
            def sciencestream():
                cut.destroy()
                lot=tki.Tk()
                               
                lot.geometry("{0}x{1}+0+0".format(lot.winfo_screenwidth(), lot.winfo_screenheight()))
                lot.title("Kendriya vidyalaya")

                b=Label(lot,text='Science Stream',font='arial 14 bold')
                myfont=Font(family="Impact",size=30)
                b.configure(font=myfont)
                b.place(x=500,y=0)

                #function for student data entry
                def detailinput():
                    import detail                   
                    

                def back4():
                    lot.destroy()
                    nxt()
                def studentdetails():
                    pii=tki.Tk()
                    pii.geometry("{0}x{1}+0+0".format(pii.winfo_screenwidth(), pii.winfo_screenheight()))
                    pii.title("Kendriya vidyalaya")
                    b=Label(pii,text='STUDENT DETAILS',font='arial 24 bold')
                    b.pack()
                    def moreinfo():
                        os.system("start EXCEL.EXE class12.xlsx")

                    def namesbio():
                        path = "marks.xlsx"# Give the location of the file
                        wb_obj = openpyxl.load_workbook(path)# workbook object is created 
                        sheet_obj = wb_obj.active 
                        m_row = sheet_obj.max_row
                        print("BIOLOGY STUDENTS:")# Loop will print all values of first column
                        for i in range(1, m_row + 1):
                            cell_obj = sheet_obj.cell(row = i, column =9)
                            cell_obj1 = sheet_obj.cell(row = i, column =3)
                            p=cell_obj.value
                            V=cell_obj1.value
                            if p=='NIL':
                                  print(V)
    
                    def namesmaths():
                        path = "G:\\Cs project\\shrey\\marks.xlsx"# Give the location of the file
                        wb_obj = openpyxl.load_workbook(path)# workbook object is created 
                        sheet_obj = wb_obj.active 
                        m_row = sheet_obj.max_row
                        print("MATHEMATICS STUDENTS:")# Loop will print all values of first column
                        for i in range(1, m_row + 1):
                            cell_obj = sheet_obj.cell(row = i, column =8)
                            cell_obj1 = sheet_obj.cell(row = i, column =3)
                            p=cell_obj.value
                            V=cell_obj1.value
                            if p=='NIL':
                                  print(V)
                                  
                    button4=Button(pii,text="  BIO STUDENTS  ",command=namesbio,width=18,height=2,bd=0,bg="DimGray",\
                       activebackground="#fff",activeforeground="#42f498",fg="#fff")
                    button4.place(x=900,y=90)
                    button4=Button(pii,text="  MATHS STUDENTS  ",command=namesmaths,width=18,height=2,bd=0,bg="DimGray",\
                       activebackground="#fff",activeforeground="#42f498",fg="#fff")
                    button4.place(x=900,y=140)                    
                    button1=Button(pii,text="    More info.    ",command=moreinfo,width=18,height=2,bd=0,bg="DodgerBlue",\
                                activebackground="#fff",activeforeground="#42f498",fg="#fff")
                    button1.place(x=600,y=600)
                        
                    button1=Button(pii,text="    Back    ",command=pii.destroy,width=18,height=2,bd=0,bg="DodgerBlue",\
                                activebackground="#fff",activeforeground="#42f498",fg="#fff")
                    button1.place(x=600,y=650)
                    

                    
                    Label(pii,text='CLASS: XII SCIENCE',font='arial 14 bold').place(x=10,y=60)
                    Label(pii,text='Class Teacher :  Mr. M Hamel',font='arial 12 bold').place(x=10,y=90)
                    Label(pii,text='total students :    30',font='arial 12 bold').place(x=10,y=110)
                    Label(pii,text='boys:               18',font='arial 12 bold').place(x=10,y=130)
                    Label(pii,text='girls:              12',font='arial 12 bold').place(x=10,y=150)
                
                    Label(pii,text='Subject opted',font='arial 14 bold').place(x=1100,y=60)
                    Label(pii,text='Maths + Computer science :  6',font='arial 12 bold').place(x=1100,y=90)
                    Label(pii,text='Maths + Hindi   :    4',font='arial 12 bold').place(x=1100,y=110)
                    Label(pii,text='Maths + Biology :    2',font='arial 12 bold').place(x=1100,y=130)
                    Label(pii,text='Biology + Hindi :   18',font='arial 12 bold').place(x=1100,y=150)
                    Label(pii,text='#The names of both bio/maths students shall be displayed on the python console.',font='arial 10 bold',bg="LavenderBlush",\
                                     activebackground="#fff",activeforeground="#42f498",fg="#ff3333").place(x=800,y=200)



                    figure2 = Figure(figsize=(4,3), dpi=100) # create a Figure 
                    subplot2 = figure2.add_subplot(111) # add a subplot
                    labels2 = 'Boys', 'Girls'
                    pieSizes = [float(18),float(12)]
                    explode2 = (0, 0.1) 
                    subplot2.pie(pieSizes, explode=explode2, labels=labels2, autopct='%1.1f%%', shadow=True, startangle=90) 
                    subplot2.axis('equal')  
                    pie2 = FigureCanvasTkAgg(figure2, pii) # create a canvas figure (matplotlib module)
                    pie2.get_tk_widget().place(x=50,y=295)
                    

                    figure = Figure(figsize=(4,3), dpi=100) # create a Figure 
                    subplot = figure.add_subplot(111) # add a subplot
                    labels = 'Maths +Cs', 'Maths + Hin ','Maths + Bio' ,'Bio + Hindi'
                    pieSizes = [float(6),float(4),float(2),float(18)]
                    explode = (0, 0,0.1,0) 
                    subplot.pie(pieSizes, explode=explode, labels=labels, autopct='%1.1f%%', shadow=True, startangle=90) 
                    subplot.axis('equal')  
                    pie = FigureCanvasTkAgg(figure, pii) # create a canvas figure (matplotlib module)
                    pie.get_tk_widget().place(x=900,y=295)
                    pii.mainloop()                  
                def marks():
                        os.system("start EXCEL.EXE excel.xlsx")

                    
                    #name=Label(k,text="student name").place(x=520,y=305)
                              
                                              
                #function closed
                button1=Button(lot,text="    Students details    ",command=studentdetails,width=18,height=2,bd=0,bg="DodgerBlue",\
                                activebackground="#fff",activeforeground="#42f498",fg="#fff")
                button1.place(x=550,y=150)
                button2=Button(lot,text="  Student data entry  ",command=detailinput,width=18,height=2,bd=0,bg="DodgerBlue",\
                                activebackground="#fff",activeforeground="#42f498",fg="#fff")
                button2.place(x=550,y=200)
                button3=Button(lot,text="          Marks                 ",command=marks,width=18,height=2,bd=0,bg="DodgerBlue",\
                                activebackground="#fff",activeforeground="#42f498",fg="#fff")
                button3.place(x=550,y=250)
                button4=Button(lot,text="     Students CCA       ",width=18,height=2,bd=0,bg="DodgerBlue",\
                                activebackground="#fff",activeforeground="#42f498",fg="#fff")
                button4.place(x=550,y=300)
                button5=Button(lot,text="          Back               ",width=18,height=2,bd=0,bg="DimGray",\
                               activebackground="#fff",activeforeground="#42f498",fg="#fff")
                button5.place(x=550,y=450)
##################################################################################################################################################################                              
            button1=Button(cut,text="    Science stream    ",command=sciencestream,width=18,height=2,bd=0,bg="DodgerBlue",\
                           activebackground="#fff",activeforeground="#42f498",fg="#fff")
            button1.place(x=550,y=150)
            button2=Button(cut,text="Commerce stream  ",width=18,height=2,bd=0,bg="DodgerBlue",\
                           activebackground="#fff",activeforeground="#42f498",fg="#fff")
            button2.place(x=550,y=200)
            button3=Button(cut,text="      Arts stream     ",width=18,height=2,bd=0,bg="DodgerBlue",\
                           activebackground="#fff",activeforeground="#42f498",fg="#fff")
            button3.place(x=550,y=250)
            def back():
                cut.destroy()
                nxt()
                
            button4=Button(cut,text="Back",command=back,width=18,height=2,bd=0,bg="DimGray",\
                       activebackground="#fff",activeforeground="#42f498",fg="#fff")
            button4.place(x=550,y=350)
            
            
        #class 12th ended    
        button7=Button(nut,text="_           CLASS XII         _",command=class12,width=18,height=2,bd=0,bg="DodgerBlue",\
                       activebackground="#fff",activeforeground="#42f498",fg="#fff")
        button7.place(x=550,y=450)


            
    #their buttons
    B1=Button(root,text="Primary classes",command=primary,width=14,height=2,bd=0,bg="DodgerBlue",\
             activebackground="#fff",activeforeground="#42f498",fg="#fff").grid(row=4,column=1)
    B2=Button(root,text="Secondary classes",command=secondary,width=14,height=2,bd=0,bg="DodgerBlue",\
             activebackground="#fff",activeforeground="#42f498",fg="#fff").grid(row=4,column=4)
    root.mainloop()
    
    
#ended
    
def insert():
    password=v1.get()
    if password=="s" :
        top.destroy()
        nxt()
    else:
        messagebox.showinfo("Login failed","wrong password")

            
B=tki.Button(top,text="Login",command=insert,width=10,height=2,bd=0,bg="DodgerBlue",\
             activebackground="#fff",activeforeground="#42f498",fg="#fff",relief=RAISED).place(x=650,y=600)

top.mainloop()
